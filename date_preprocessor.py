import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook

class DatePreprocessorGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Преобразователь дат Excel")
        self.root.geometry("400x200")

        # Кнопка для выбора входного файла
        self.select_button = tk.Button(root, text="Выбрать Excel-файл", command=self.select_file)
        self.select_button.pack(pady=10)

        # Метка для отображения пути к файлу
        self.file_path_label = tk.Label(root, text="Файл не выбран")
        self.file_path_label.pack(pady=5)

        # Кнопка для сохранения обработанного файла
        self.save_button = tk.Button(root, text="Сохранить обработанный файл", command=self.save_file, state='disabled')
        self.save_button.pack(pady=10)

    def select_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            self.file_path = file_path
            self.file_path_label.config(text=f"Выбран файл: {file_path}")
            self.save_button.config(state='normal')
            messagebox.showinfo("Успех", "Файл выбран. Нажмите 'Сохранить обработанный файл' для продолжения.")

    def save_file(self):
        if not hasattr(self, 'file_path'):
            messagebox.showerror("Ошибка", "Сначала выберите файл!")
            return

        try:
            # Читаем Excel-файл
            df = pd.read_excel(self.file_path)

            # Отладочное сообщение: показываем исходный формат столбца "По дням"
            print("Исходный формат столбца 'По дням':")
            print(df.iloc[:, 1].head())

            # Преобразуем столбец "По дням" (второй столбец, индекс 1) в формат даты
            df.iloc[:, 1] = pd.to_datetime(df.iloc[:, 1], format='%Y-%m-%d %H:%M:%S', errors='coerce')

            # Отладочное сообщение: показываем формат даты после преобразования
            print("Формат даты после преобразования в datetime:")
            print(df.iloc[:, 1].head())

            # Здесь оставляем данные как datetime, чтобы Excel применял форматирование,
            # которое потом изменим на текстовый.

            # Открываем диалоговое окно для сохранения файла
            output_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx *.xls")],
                title="Сохранить обработанный файл"
            )

            if output_path:
                # Сохраняем обработанный файл через ExcelWriter без указания datetime_format
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)

                # Открываем файл через openpyxl для изменения формата ячеек столбца "По дням"
                wb = load_workbook(output_path)
                ws = wb.active

                # Предполагаем, что столбец "По дням" находится во втором столбце (B)
                for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
                    for cell in row:
                        # Если значение ячейки является датой, преобразуем его в строку в нужном формате
                        if cell.value is not None and hasattr(cell.value, 'strftime'):
                            cell.value = cell.value.strftime('%Y-%m-%d')
                        else:
                            cell.value = str(cell.value)
                        # Устанавливаем текстовый формат ячейки
                        cell.number_format = '@'
                wb.save(output_path)

                messagebox.showinfo("Успех", f"Обработанный файл сохранен как: {output_path}")
                self.save_button.config(state='disabled')
                self.file_path_label.config(text="Файл не выбран")
            else:
                messagebox.showwarning("Предупреждение", "Сохранение отменено.")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка при обработке файла: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = DatePreprocessorGUI(root)
    root.mainloop()
